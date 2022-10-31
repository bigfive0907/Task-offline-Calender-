import tkinter as tk
from tkinter import GROOVE, RAISED, RIDGE, SUNKEN, ttk
import datetime as dt
import tkinter
from turtle import color
from unittest import result
import openpyxl as px
import webbrowser as web
from openpyxl.styles import Alignment


#使いやすいデザインに変える

# 日付と予定をその人のシートのエクセルに書き込む（別ウインドウ）
# 日付が当日ならば予定を表示する（シートの列でforでデータを持ってくる）
class Application(tk.Frame):
    def __init__(self,master=None):
        
        dt_now = dt.datetime.now()
        self.dates = ['','','','','','','','','','','','','','','','']
        for i in range(-10,10):
            self.dates[i] = dt_now + dt.timedelta(days=i)
        self.month_today = str(dt_now.month)
        self.day_today = str(dt_now.day)
        self.today = dt_now.strftime('%Y年%m月%d日')
        self.year = 2022
        # Google APIの準備をする

        tk.Frame.__init__(self, master,) #初期化
        #self.pack()
        bookpath = r"excel\schedule.xlsx"
        book = px.load_workbook(bookpath)
        ws = book.active
        self.master.geometry("500x400")
        self.master.title("Task")
        
        self.menu_bar()
        self.widgets()
        self.pack()

    def menu_bar(self):
        men = tk.Menu(self)
        self.master.config(menu = men)

        menu_file = tkinter.Menu(self) 
        men.add_cascade(label='menu', menu=menu_file)

        #親メニューに子メニュー（開く・閉じる）を追加する 
        menu_file.add_command(label='スケジュールを編集',command=self.input_schedule) 
        #menu_file.add_command(label='一週間の予定を表示',command=self.sch_open )
        #menu_file.add_command(label='アラーム', )
        menu_file.add_command(label='Googleカレンダーを開く', command=self.google_open)
        menu_file.add_command(label='終了',command=self.close_window)

    def widgets(self):
        result = self.text_open()
        week_result = self.sch_open()
    
        self.today_label = ttk.Label(self)
        self.today_label.configure(text=(self.today +"の予定"), font=("メイリオ", "8"))
        self.today_label.pack(side = tk.TOP)

        self.today_text = tk.StringVar(self)
        self.today_text.set(''.join(result))
        self.today_label = ttk.Label(self, textvariable = self.today_text, font=("メイリオ", "12", "bold"))
        self.today_label.pack(side = tk.TOP)

        self.week_label = ttk.Label(self)
        self.week_label.configure(text=("今後の予定"), font=("メイリオ", "8"))
        self.week_label.pack(side = tk.TOP)

        self.week1_text = tk.StringVar(self)
        self.week1_text.set(''.join(week_result))
        self.week1_label = ttk.Label(self, textvariable = self.week1_text, font=("メイリオ", "12", "bold"))
        self.week1_label.pack(side = tk.TOP)

        
    def google_open(self):
        url="https://calendar.google.com/calendar/u/0/r?tab=mc"
        web.open(url,new=1,autoraise=True)
    
    def sch_open(self):
        bookpath = r"excel\schedule.xlsx" #ok
        book = px.load_workbook(bookpath)
        ws = book.active
        self.week_result = []

        for j in range(0,9):
            
            for i in range(2,ws.max_row+1):

                sch_month = str(ws.cell(i,1).value)
                sch_day = str(ws.cell(i,2).value)
                other_month = str(self.dates[j].month)
                other_day = str(self.dates[j].day)
                    
                if(other_month == sch_month  and  other_day == sch_day):
                    sch_month = str(ws.cell(i,1).value)
                    sch_day = str(ws.cell(i,2).value)
                    sch_hour = str(ws.cell(i,3).value)
                    sch_time = str(ws.cell(i,4).value)
                    sch_main = str(ws.cell(i,5).value)
                    sch_id = str(ws.cell(i,6).row)
                    sch_comment = str(ws.cell(i,6).value)
                    week_text = ( 'ID: ' + sch_id + '     ' + sch_month +'月'+ sch_day +'日'+'    '+ sch_hour + '時' + sch_time + '分' + '    ' + sch_main +'   '+ sch_comment + '\n')
                    self.week_result.append(week_text)
        return self.week_result
    
    def text_open(self):
        
        bookpath = r"excel\schedule.xlsx" #ok
        book = px.load_workbook(bookpath)
        ws = book.active
        result = []
        #何もない行を削除
        for i in range(2,ws.max_row+1):
            if ws.cell(i, 1).value== None:
                ws.delete_rows(i)
        book.save(bookpath)

        for i in range(2,ws.max_row+1):

            sch_month = str(ws.cell(i,1).value)
            sch_day = str(ws.cell(i,2).value)
            
                
            if(self.month_today == sch_month  and self.day_today == sch_day):
                sch_hour = str(ws.cell(i,3).value)
                sch_time = str(ws.cell(i,4).value)
                sch_main = str(ws.cell(i,5).value)
                sch_id = str(ws.cell(i,6).row)
                sch_comment = str(ws.cell(i,6).value)
                text = ( 'ID: ' + sch_id + '     ' + sch_hour + '時' + sch_time + '分' + '  ' + sch_main + '   '+ sch_comment +'\n')
                result.append(text)
        return result

        
    def book_save(self):
        bookpath = r"excel\schedule.xlsx"
        book = px.load_workbook(bookpath)
        ws = book.active
        data = [0,0,0,0,0,0]
        #データを取得
        data[0] = self.month_combo.get()
        data[1] = self.day_combo.get()
        data[2] = self.hour_combo.get()
        data[3] = self.time_combo.get()
        data[4] = self.entry_event.get()
        data[5] = self.entry_comment.get()
        #データの書き込み
        ws.insert_rows(2)#書き込むために行を一つ用意
        for i, datas in enumerate(data):
            ws.cell(row=2, column=1+i, value=data[i])

        book.save(bookpath)
        result = self.text_open()
        week_result = self.sch_open()
        self.today_text.set(''.join(result))
        self.week1_text.set(''.join(week_result))
        self.save_text.set(data[4] + 'の予定が'+ data[0]+'月'+data[1]+'日'+data[2]+'時'+data[3]+'分に登録されました')
    

    def clear(self):
        bookpath = r"excel\schedule.xlsx"
        book = px.load_workbook(bookpath)
        ws = book.active

        self.clear_id = self.entry_comment.get()

        for i in range(2,ws.max_row+1):
            sch_id = str(ws.cell(i,6).row)
            if(self.clear_id == sch_id):
                ws.delete_rows(i)

        book.save(bookpath)
        result = self.text_open()
        week_result = self.sch_open()
        self.today_text.set(''.join(result))
        self.week1_text.set(''.join(week_result))



    def close_window(self):
        self.master.destroy()

    def input_schedule(self): #sub_window

        sub_window = tk.Toplevel()
        sub_window.geometry("300x200")

        self.label_month = tk.Label(sub_window)
        self.label_month.configure(text = "月")
        self.label_month.place(x=10,y=10)

        self.month = tk.StringVar()
        month_value = ("1","2","3","4","5","6","7","8","9","10","11","12")
        self.month_combo = ttk.Combobox(sub_window,values = month_value)
        self.month_combo.place(x=80,y=10)


        self.label_day = tk.Label(sub_window)
        self.label_day.configure(text = "日")
        self.label_day.place(x=10,y=30)

        self.day = tk.StringVar()
        day_value = ("1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31")
        self.day_combo = ttk.Combobox(sub_window,values = day_value)
        self.day_combo.place(x=80,y=30)


        self.label_hour = tk.Label(sub_window)
        self.label_hour.configure(text = "時")
        self.label_hour.place(x=10,y=50)

        hour_value = ("0","1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23")
        self.hour_combo = ttk.Combobox(sub_window,values = hour_value)
        self.hour_combo.place(x=80,y=50)

        self.label_time = tk.Label(sub_window)
        self.label_time.configure(text = "分")
        self.label_time.place(x=10,y=70)

        self.time = tk.StringVar()
        time_value = ("0","1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59")
        self.time_combo = ttk.Combobox(sub_window,values = time_value)
        self.time_combo.place(x=80,y=70)
     
        self.label_time = tk.Label(sub_window)
        self.label_time.configure(text = "内容")
        self.label_time.place(x=10,y=90)

        self.event = tk.StringVar()
        self.entry_event = ttk.Entry(sub_window)
        self.entry_event.configure(textvariable = self.event)
        self.entry_event.place(x=80,y=90)

        self.label_comment = tk.Label(sub_window)
        self.label_comment.configure(text = "コメント(ID)")
        self.label_comment.place(x=10,y=110)

        self.comment = tk.StringVar()
        self.entry_comment = ttk.Entry(sub_window)
        self.entry_comment.configure(textvariable = self.comment)
        self.entry_comment.place(x=80,y=110)


        self.save_button = ttk.Button(sub_window)
        self.save_button.configure(text="Save")
        self.save_button.configure(command = self.book_save)
        self.save_button.place(x=20, y=140)
  
        self.clear_button = ttk.Button(sub_window)
        self.clear_button.configure(text="Clear(ID)")
        self.clear_button.configure(command = self.clear)
        self.clear_button.place(x=120, y=140)
        
        self.save_text = tk.StringVar()
        self.save_text.set('')
        self.add_save_label = ttk.Label(sub_window,textvariable = self.save_text, font=("メイリオ", "7","bold"))
        self.add_save_label.place(x=20,y=180)

        
def main():
    root = tk.Tk()
    #root.overrideredirect(1) #ボーダーレス
    app = Application(master=root)
    app.mainloop()

if __name__ == "__main__":
    main()



   



